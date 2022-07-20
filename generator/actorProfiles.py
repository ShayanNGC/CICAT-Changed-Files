# This file was created by Shayan Bathaee on 7/20/2022

# USAGE: 
# in a command prompt, run the following command:
# python scenActors.py <threat actor folder name> <scenario spreadsheet name>
# NOTE: In some systems, the excel file must be closed for this program to work

# This program takes two inputs. The first input is a folder path. The folder should contain one or more JSON files 
# with information about a threat actor and the threats they use. The JSON files come from MITRE's
# ATT&CK navigator. The second input to this program is a scenario spreadsheet. This scenario spreadsheet needs to have a sheet 
# titled "ACTOR PROFILES". The sheet may or may not be empty. 

# If the user has provided valid inputs, this program reads through the JSON file and updates the scenario spreadsheet with the
# required threat actor information. The purpose of this program is to automate the task of filling in the "ACTOR PROFILES" sheet
# of the scenario spreadsheet.

from openpyxl import Workbook, load_workbook
import os
import sys
import json

# get file names
folder_name = sys.argv[1]
spreadsheet_name = sys.argv[2]

# open the workbook and worksheet
wb = load_workbook(spreadsheet_name)
ws = wb["ACTOR PROFILES"]

# Clear the excel file
ws.delete_rows(1, ws.max_row + 1)

# Write the column names
# write the master cell
ws.cell(row=1, column=1).value = "MASTER"

# write the threat actor names
col = 2
actor_column = {}                                                       # dictionary, key = threat actor, value = column they are in
for filename in os.listdir(os.getcwd() + "\\" + folder_name):           # for every file name in the folder
    with open(os.path.join(os.getcwd(), folder_name, filename)) as f:   # open the file and load the json data
        data = json.load(f)
        ws.cell(row=1, column=col).value = data["name"]                 # write the actor name to the next column cell
        actor_column[data["name"]] = col                                # insert actor_column data
        col += 1
    f.close()                                                           # close the file

# Create a dictionary of the threat actor data
technique_actors = {}                                                   # dictionary, key = technique, value = list of actors using that technique
for filename in os.listdir(os.getcwd() + "\\" + folder_name):           # for every file name in the folder
    with open(os.path.join(os.getcwd(), folder_name, filename)) as f:   # open the file and load the json data
        data = json.load(f)
        for technique in data["techniques"]:                            # for every technique in the file
            techniqueID = technique["techniqueID"]
            if techniqueID in technique_actors:                         # if our dictionary already has the technique, just add the actor name
                technique_actors[techniqueID].add(data["name"])
            else:                                                       # if our dictionary doesn't have the technique
                technique_actors[techniqueID] = {data["name"]}          # create a new set with one element, the actor name, as the key

# Write the dictionary data to the spreadsheet
row_index = 2
for t in technique_actors:                                              # for every technique
    ws.cell(row=row_index, column=1).value = t                          # write the technique in the 'master' column
    for a in technique_actors[t]:                                       # for every actor that uses that technique
        ws.cell(row=row_index, column=actor_column[a]).value = 'X'      # write an X int the actor's column
    row_index += 1

# Save the worksheet
wb.save(spreadsheet_name)
