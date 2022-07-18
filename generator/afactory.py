# -*- coding: utf-8 -*-
"""
::::::::::::::::::::::::  Critical Infrastructure Cyberspace Analysis Tool (CICAT)  :::::::::::::::::::::::::::::::::::::::

                                            NOTICE
                                            
The contents of this material reflect the views of the author and/or the Director of the Center for Advanced Aviation 
System Development (CAASD), and do not necessarily reflect the views of the Federal Aviation Administration (FAA) 
or the Department of Transportation (DOT). Neither the FAA nor the DOT makes any warranty or guarantee, or promise, 
expressed or implied, concerning the content or accuracy of the views expressed herein. 

This is the copyright work of The MITRE Corporation and was produced for the U.S. Government under Contract Number 
DTFAWA-10-C-00080 and is subject to Federal Aviation Administration Acquisition Management System Clause 3.5-13, 
Rights in Data-General, Alt. III and Alt. IV (Oct. 1996). No other use other than that granted to the U.S. Government, 
or to those acting on behalf of the U.S. Government, under that Clause is authorized without the express written permission 
of The MITRE Corporation. For further information, please contact The MITRE Corporation, Contract Office, 7515 Colshire Drive, 
McLean, VA 22102 (703) 983-6000. ©2020 The MITRE Corporation. 

The Government retains a nonexclusive, royalty-free right to publish  or reproduce this document, or to allow others to do so, for 
“Government Purposes Only.”                                           
                                            
(c) 2020 The MITRE Corporation. All Rights Reserved.

::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
afactory.py - Factory class for ATT&CK data
::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
"""

from matplotlib.pyplot import xlabel
import openpyxl
import json
import os

from sqlalchemy import false

from amodel import TTP
from amodel import MIT
from amodel import MALWARE
from amodel import THREATACTOR
from amodel import CYBERTOOL
from amodel import ATKRELATION
from columnloader import LoadColsInSpreadsheet, LoadColNames


def isCOA (entry):
    if not(entry.find ('course-of-action') < 0):
        return True
    return False

def isTTP(entry):
    if not(entry.find ('attack-pattern') < 0):
        return True
    return False

def isMAL(entry):
    if not(entry.find ('malware') < 0):
        return True
    return False

def isACT(entry):
    if not(entry.find('intrusion-set') < 0):
        return True
    return False

def isTOOL(entry):
    if not(entry.find('tool--') < 0):
        return True
    return False

def findOBJECT(pattern, listx):
    for t in listx:
        if (t.getID() == pattern):
            return t

def aslist (val, delimiter):
    if not(val) or val == 'None':
        return None
    elif val == '' or val == ' ' or val == 'undefined':
        return []
    elif delimiter in val: 
        return val.split(delimiter)
    else:
        return [val]

class ATTACK_FACTORY():

    # LOAD OPTIONS FOR ATTACK DATA: STIX, JSON, SPREAD, SQL
    
        def __init__ (self, loadOPT, fname, trace ):           
          if trace:
              print ('ATT&CK factory constructed..')
          self.trace = trace
          self.fname = fname
          self.loadFromSpreadsheet = False
             
          if loadOPT == 'JSON':
             if self.trace:
                 print('Opening local json dataset')
             # added encoding='utf-8' to account for bytes that couldn't be read (namely ”)
             with open(os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'data', 'ATK', 'attack.json')), encoding='utf-8') as data_file:    
               self.all_attack = json.load(data_file)

          elif loadOPT == 'STIX':
             if self.trace:
                 print ('Connecting to STIX/TAXII service')
             try:
                from attackcti import attack_client
                self.lift = attack_client()            
                self.all_attack = self.lift.get_all_stix_objects()
             except:
                print ('Cannot connect to STIX/TAXII service: ')
                print ('Most likely cause: application cannot reach the Internet')
                raise

          elif loadOPT == 'SPREAD':
              if self.trace:
                  print ('Loading data from spreadsheet', self.fname)
              self.loadFromSpreadsheet = True

          else:
              if self.trace:
                 print ('WARNING! ATTACK factory: Unsupported load option:', loadOPT)

          self.ttps = []
          self.groups = None #self.all_attack['groups']
          self.malwares = None #self.all_attack['malware']
          self.mitigations = None #self.all_attack['mitigations']
          self.techniques = None #self.all_attack['techniques']
          self.tools = None #self.all_attack['tools']  
          self.relationships = None #self.all_attack['relationships']    
          self.groupprofiles = None
               
        def loadGroupsFromJSON (self):
           ret = []
           self.groups = self.all_attack['groups']
           
           for group in self.groups:
                for x in group['objects']: # new format puts all of the group's data in the object key
                    external_references_list = []
                    if 'external_references' in x:
                        # get the group's url, matrix, and group id (first element's key-value pairs in external references)
                        ref0 = x['external_references'][0]
                        url = ref0['url'] if 'url' in ref0 else None
                        matrix = ref0['source_name'] if 'source_name' in ref0 else None
                        group_id = ref0['external_id'] if 'external_id' in ref0 else None

                        # get external reference data
                        for ref in x['external_references']:
                            # assign variables
                            url = ref['url'] if 'url' in ref else None
                            source_name = ref['source_name'] if 'source_name' in ref else None
                            description = ref['description'] if 'description' in ref else None
                            # build reference (add url, source name, and description in that order if available)
                            ref_string = ""
                            if url:
                                ref_string += url
                            if source_name:
                                if ref_string != "":
                                    ref_string += ", "
                                ref_string += source_name
                            if description:
                                if ref_string != "":
                                    ref_string += ', '
                                ref_string += description
                            # if there was data in the reference, add it
                            if ref_string != "":
                                external_references_list.append(ref_string)

                    # append existing data. append None if the key does not exist
                    ret.append(THREATACTOR (x['created'] if 'created' in x else None,
                                    x['created_by_ref'] if 'created_by_ref' in x else None, 
                                    x['name'] if 'name' in x else None,
                                    x['aliases'] if 'aliases' in x else None, 
                                    x['description'] if 'description' in x else None,
                                    group_id,                   # the group id 
                                    external_references_list,   # the external reference data determined earlier
                                    x['id'] if 'id' in x else None, 
                                    matrix,                     # the matrix is in the first external reference
                                    x['modified'] if 'modified' in x else None,
                                    x['type'] if 'type' in x else None, 
                                    url))                       # the url of the group
           
           return ret


        def loadGroupsFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows:
              ret.append(THREATACTOR (row[0].value,
                                      row[1].value,
                                      row[2].value,
                                      (row[3].value).split(),
                                      row[4].value,
                                      row[5].value,
                                      (row[6].value).split(),
                                      row[7].value,
                                      row[8].value,
                                      row[9].value,
                                      row[10].value,
                                      row[11].value ))
              
          del (ret[0])
          return ret
      
        def loadProfileNames(self, fname):
            return LoadColNames (fname, 'ACTOR PROFILES')

        # group profile are loaded from the ACTOR PROFILES tab in the INFRASTRUCTURE MODEL spreadsheet         
        def loadGroupProfile (self, dataset, fname, pname):
            # Profiles are created one at a time with self.groupprofiles  used as a cache for actor profile data.  
            if not (self.groupprofiles):
               self.groupprofiles =  LoadColsInSpreadsheet (fname, 'ACTOR PROFILES')
               
            if not (self.groupprofiles):
                if self.trace:
                    print ('WARNING! cannot load group profile data')
                return
                    
            ttplist = self.groupprofiles[pname]
            if not (ttplist):
                if self.trace:
                    print ('WARNING! cannot find group profile:', pname)
                return
                      
            ret = THREATACTOR(None, None, pname, pname, "", pname, None, 'intrustion-set-'+pname, 'mitre-attack', None, 'intrusion set', None )

            for j in self.groupprofiles[pname]:
                bfound = False
                for a in dataset['ATT&CK']:
                    if a.getTECHID() == j:
                        ret.addUses( a, None)
                        bfound = True
                        break

                if not(bfound):
                    for i in dataset['ATK4ICS TTPs']:
                       if i.getTECHID() == j:
                          ret.addUses( i, None)
                          bfound = True
                          break                        
               
                if not(bfound):
                    if self.trace:
                        print ('WARNING! TTP', j, 'not found')                

            return ret

        def loadGroups(self):
            if self.trace:
                print ('Loading ATT&CK Groups data..')

            if self.loadFromSpreadsheet:
                return self.loadGroupsFromSheet (self.fname, 'ATKGROUPS')
            else:
                return self.loadGroupsFromJSON ()

        def loadMalwaresFromJSON (self):
            ret = []
            self.malwares = self.all_attack['malware']
            
            for malware in self.malwares:
                for x in malware['objects']:
                    external_references_list = []
                    if 'external_references' in x:
                        # get the malware's url, matrix, and group id (first element's key-value pairs in external references)
                        ref0 = x['external_references'][0]
                        url = ref0['url'] if 'url' in ref0 else None
                        matrix = ref0['source_name'] if 'source_name' in ref0 else None
                        software_id = ref0['external_id'] if 'external_id' in ref0 else None

                        # get external reference data
                        for ref in x['external_references']:
                            # assign variables
                            url = ref['url'] if 'url' in ref else None
                            source_name = ref['source_name'] if 'source_name' in ref else None
                            description = ref['description'] if 'description' in ref else None
                            # build reference (add url, source name, and description in that order if available)
                            ref_string = ""
                            if url:
                                ref_string += url
                            if source_name:
                                if ref_string != "":
                                    ref_string += ", "
                                ref_string += source_name
                            if description:
                                if ref_string != "":
                                    ref_string += ', '
                                ref_string += description
                            # if there was data in the reference, add it
                            if ref_string != "":
                                external_references_list.append(ref_string)
                    
                    ret.append (MALWARE (x['created'] if 'created' in x else None,
                                        x['created_by_ref'] if 'created_by_ref' in x else None,
                                        x['id'] if 'id' in x else None,
                                        matrix,
                                        x['modified'] if 'modified' in x else None,
                                        x['name'] if 'name' in x else None, 
                                        x['x_mitre_aliases'] if 'x_mitre_aliases' in x else None, 
                                        x['description'] if 'description' in x else None,
                                        software_id, 
                                        x['labels'] if 'labels' in x else None,
                                        x['x_mitre_platforms'] if 'x_mitre_platforms' in x else None, 
                                        external_references_list, 
                                        x['type'] if 'type' in x else None, 
                                        url))
            
            return ret
        
        def loadMalwaresFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows:            
                 ret.append (MALWARE (row[0].value,
                                      row[1].value,
                                      row[2].value,
                                      row[3].value,
                                      row[4].value,
                                      row[5].value,
                                      (row[6].value).split(','),
                                      row[7].value,
                                      row[8].value,
                                      (row[9].value).split(),
                                      (row[10].value).split(),
                                      (row[11].value).split(),
                                      row[12].value,
                                      row[13].value ))
          del (ret[0])
          return ret


        def loadMalwares(self):
            if self.trace:
                print ('Loading ATT&CK Malware data..')
                
            if self.loadFromSpreadsheet:
                return self.loadMalwaresFromSheet (self.fname, 'ATKMALWARE')
            else:
                return self.loadMalwaresFromJSON ()


        def loadMitigationsFromJSON(self):
            ret = []
            self.mitigations = self.all_attack['mitigations']
            
            for mitigation in self.mitigations:
                for x in mitigation['objects']:
                    external_references_list = []
                    if 'external_references' in x:
                        ref0 = x['external_references'][0]
                        url = ref0['url'] if 'url' in ref0 else None
                        matrix = ref0['source_name'] if 'source_name' in ref0 else None
                        technique_id = ref0['external_id'] if 'external_id' in ref0 else None

                        # determine the external reference data (url, source name, description)
                        ref_string = ""
                        for ref in x['external_references']:
                            # assign variables
                            url = ref['url'] if 'url' in ref else None
                            source_name = ref['source_name'] if 'source_name' in ref else None
                            description = ref['description'] if 'description' in ref else None
                            # build reference (add url, source name, and description in that order if available)
                            if source_name:
                                ref_string += source_name
                            if url:
                                if source_name:
                                    ref_string += ", "
                                ref_string += url
                            if description:
                                if url or source_name:
                                    ref_string += ', '
                                ref_string += description
                            # if there was data in the reference, add the end
                            if url or source_name or description:
                                ref_string += "; \n"
                        if ref_string != "":
                            ref_string = ref_string[:-1]
                            external_references_list.append(ref_string)         
                    
                    ret.append (MIT (x['created'] if 'created' in x else None,
                                           x['created_by_ref'] if 'created_by_ref' in x else None,
                                           x['id'] if 'id' in x else None,
                                           matrix,
                                           # couldn't find 'mitigation' in new file. Replaced with 'name'
                                           x['name'] if 'name' in x else None,
                                           x['description'] if 'description' in x else None,
                                           external_references_list,
                                           x['modified'] if 'modified' in x else None,
                                           technique_id,
                                           x['type'] if 'type' in x else None,
                                           url ))
            
            return ret

        def loadMitigationsFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []       
          for row in sheet.rows:                             
              r0 = row[0].value
              r1 = row[1].value
              r2 = row[2].value
              r3 = row[3].value
              r4 = row[4].value
              r5 = row[5].value
              r6 = aslist(row[6].value, ' ')
              r7 = row[7].value
              r8 = row[8].value
              r9 = row[9].value
              r10 = row[10].value                          
              ret.append (MIT (r0, r1, r2, r3, r4, r5, r6, r7, r8, r9, r10 ))
                                                 
          del (ret[0])
          return ret

        def loadMitigations(self):
            if self.trace:
                print ('Loading ATT&CK Mitigations data..')
                
            if self.loadFromSpreadsheet:
                return self.loadMitigationsFromSheet (self.fname, 'ATKMITIGATION')
            else:
                return self.loadMitigationsFromJSON ()

        def loadTechniques(self):
            if self.trace:
                print ('Loading ATT&CK TTP data..')
                
            if self.loadFromSpreadsheet:
                return self.loadTechniquesFromSheet (self.fname, 'ATT&CK')
            else:
                return self.loadTechniquesFromJSON ()

        def loadTechniquesFromJSON(self):
             ret = []
             self.techniques = self.all_attack['techniques']
             
             for technique in self.techniques:
                for x in technique['objects']:
                    external_references_list = []
                    capec_id_list = []
                    capec_url_list = []
                    if 'external_references' in x:
                        ref0 = x['external_references'][0]
                        
                        # get all external reference data
                        ref_string = ""
                        for ref in x['external_references']:
                            # if we are at the first element, get matrix and technique id
                            if ref == ref0:
                                matrix = ref0['source_name'] if 'source_name' in ref0 else None
                                technique_id = ref0['external_id'] if 'external_id' in ref0 else None
                                
                            # if the element is a capec, get the capec id and url
                            elif 'source_name' in ref and ref['source_name'] == 'capec':
                                capec_id = ref['external_id'] if 'external_id' in ref else None
                                capec_url = ref['url'] if 'url' in ref else None
                                capec_id_list.append(capec_id)
                                capec_url_list.append(capec_url)
                            
                            # check for errors and add the reference (even if it is ref0 or capec)
                            # assign variables
                            url = ref['url'] if 'url' in ref else None
                            source_name = ref['source_name'] if 'source_name' in ref else None
                            description = ref['description'] if 'description' in ref else None
                            # build reference (add url, source name, and description in that order if available)
                            if source_name:
                                ref_string += source_name
                            if url:
                                if source_name:
                                    ref_string += ', '
                                ref_string += url
                            if description:
                                if source_name or url:
                                    ref_string += ", "
                                ref_string += description
                            # if there was data in the reference, add it to the reference string
                            if url or source_name or description:
                                ref_string += "; \n"
                        if ref_string != "":
                            ref_string = ref_string[:-1]
                            external_references_list.append(ref_string)


                    if 'kill_chain_phases' in x:
                        # get all tactics
                        tactics_list = []
                        for phase in x['kill_chain_phases']:
                            tactic = phase['phase_name'] if 'phase_name' in phase else None
                            tactics_list.append(tactic)
                    
                    ret.append(TTP (capec_id_list,
                                        capec_url_list, 
                                        x['x_mitre_contributors'] if 'x_mitre_contributors' in x else None,
                                        x['created'] if 'created' in x else None,
                                        x['created_by_ref'] if 'created_by_ref' in x else None,
                                        x['x_mitre_data_sources'] if 'x_mitre_data_sources' in x else None,
                                        x['x_mitre_defense_bypassed'] if 'x_mitre_defence_bypassed' in x else None,
                                        # These next 4 don't show up in the new ATT&CK file
                                        x['detectable_by_common_defenses'] if 'detectable_by_common_defenses' in x else None,
                                        x['detectable_explanation'] if 'detectable_explanation' in x else None,
                                        x['difficulty_explanation'] if 'difficulty_explanation' in x else None,
                                        x['difficulty_for_adversary'] if 'difficulty_for_adversary' in x else None,
                                        x['x_mitre_effective_permissions'] if 'x_mitre_effective_permissions' in x else None,
                                        x['id'],
                                        matrix,
                                        x['modified'],
                                        # can't find network requirements in new file
                                        x['network_requirements'] if 'network_requirements' in x else None,
                                        x['object_marking_refs'] if 'object_marking_refs' in x else None,
                                        x['x_mitre_permissions_required'] if 'x_mitre_permissions_required' in x else None,
                                        x['x_mitre_platforms'] if 'x_mitre_platforms' in x else None,
                                        x['x_mitre_remote_support'] if 'x_mitre_remote_support' in x else None,
                                        x['x_mitre_system_requirements'] if 'x_mitre_system_requirements' in x else None,
                                        tactics_list,
                                        # can't find tactic type in new file
                                        x['tactic_type'] if 'tactic_type' in x else None,
                                        x['name'] if 'name' in x else None,
                                        x['description'] if 'description' in x else None,
                                        x['x_mitre_detection'] if 'x_mitre_detection' in x else None,
                                        technique_id,
                                        external_references_list,
                                        x['type'] if 'type' in x else None)) 
             
             self.ttps = ret
             return ret

        def loadTechniquesFromSheet (self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows:  
              
              r0 = aslist(row[0].value, ' ')
              r1 = aslist(row[1].value, ' ')
              r2 = aslist(row[2].value, ',')
              r3 = row[3].value
              r4 = row[4].value
              r5 = aslist(row[5].value, ',')
              r6 = aslist(row[6].value, ' ')
              r7 = row[7].value
              r8 = row[8].value
              r9 = row[9].value
              r10 = row[10].value
              r11 = aslist(row[11].value, ' ')
              r12 = row[12].value
              r13 = row[13].value
              r14 = row[14].value              
              r15 = row[15].value               
              r16 = aslist(row[16].value, ' ')              
              r17 = aslist(row[17].value, ' ')               
              r18 = aslist(row[18].value, ' ') 
              r19 = row[19].value 
              r20 = aslist(row[20].value, ' ')
              r21 = aslist(row[21].value, ' ') 
              r22 = row[22].value 
              r23 = row[23].value 
              r24 = row[24].value 
              r25 = row[25].value 
              r26 = row[26].value 
              r27 = aslist(row[27].value, ' ') 
              r28 = row[28].value               
                           
              ret.append (TTP (r0, r1, r2, r3, r4, r5, r6, r7, r8, r9, r10,
                               r11, r12, r13, r14, r15, r16, r17, r18, r19,
                               r20, r21, r22, r23, r24, r25, r26, r27, r28 ) )
         
          del (ret[0])
          return ret             
             

        def loadTools(self):
            if self.trace:
                print ('Loading ATT&CK Tools data..')
            if self.loadFromSpreadsheet:
                return self.loadToolsFromSheet (self.fname, 'ATKTOOL' )
            else:
                return self.loadToolsFromJSON ( )
            
        def loadToolsFromJSON(self):
            ret = []
            self.tools = self.all_attack['tools']
            
            for tool in self.tools:
                for x in tool['objects']:
                    external_references_list = []
                    if 'external_references' in x:
                        ref0 = x['external_references'][0]
                        matrix = ref0['source_name'] if 'source_name' in ref0 else None
                        url = ref0['url'] if 'url' in ref0 else None
                        software_id = ref0['external_id'] if 'external_id' in ref0 else None
                        for ref in x['external_references']:
                            # assign variables
                            url = ref['url'] if 'url' in ref else None
                            source_name = ref['source_name'] if 'source_name' in ref else None
                            description = ref['description'] if 'description' in ref else None
                            # build reference (add url, source name, and description in that order if available)
                            ref_string = ""
                            if url:
                                ref_string += url
                            if source_name:
                                if ref_string != "":
                                    ref_string += ", "
                                ref_string += source_name
                            if description:
                                if ref_string != "":
                                    ref_string += ', '
                                ref_string += description
                            # if there was data in the reference, add it
                            if ref_string != "":
                                external_references_list.append(ref_string)
                    
                    ret.append(CYBERTOOL (x['created'] if 'created' in x else None,
                                    x['created_by_ref'] if 'created_by_ref' in x else None,
                                    x['id'] if 'id' in x else None,
                                    matrix,
                                    x['modified'] if 'modified' in x else None,
                                    x['name'] if 'name' in x else None,
                                    x['x_mitre_aliases'] if 'x_mitre_aliases' in x else None,
                                    x['description'] if 'description' in x else None,
                                    software_id,
                                    x['labels'] if 'labels' in x else None,
                                    x['x_mitre_platforms'] if 'x_mitre_platforms' in x else None,
                                    external_references_list,
                                    x['type'] if 'type' in x else None, 
                                    url  )) 
            
            return ret             


        def loadToolsFromSheet(self, fname, sname):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]
          ret = []         
          for row in sheet.rows: 
                    ret.append(CYBERTOOL (row[0].value,
                                      row[1].value,
                                      row[2].value,
                                      row[3].value,
                                      row[4].value,
                                      row[5].value,
                                      (row[6].value).split(','),
                                      row[7].value,
                                      row[8].value,
                                      (row[9].value).split(),
                                      (row[10].value).split(),
                                      (row[11].value).split(),
                                      row[12].value,
                                      row[13].value ))     
          del (ret[0])
          return ret             


        def findTTP(self, ID, dataset):
#            if not(self.ttps):
#                print ('No TTPs to find.' )
#                return            
            for j in dataset['ATT&CK']:
                if (ID == j.getTECHID()):
                    return j                        


        def loadTTPExtension (self, dataset, filename, sheetname ):   
          if self.trace:
              print ('Loading ATT&CK extensions:', sheetname)
          book = openpyxl.load_workbook(filename, data_only=True) 
          sheet = book[sheetname]       
          for row in sheet.rows:
            ttp = self.findTTP(row[0].value, dataset)
            if (ttp):
                ttp.setP(row[1].value )
                       

        def find (self, pat, myDATAWARE ):
            if isCOA(pat):
               return findOBJECT(pat, myDATAWARE['ATKMITIGATION'] )
            elif isTTP(pat):
               return findOBJECT(pat, myDATAWARE['ATT&CK'])
            elif isACT(pat):
               return findOBJECT(pat, myDATAWARE['ATKGROUPS'])
            elif isMAL(pat):
               return findOBJECT (pat, myDATAWARE['ATKMALWARE'])
            elif isTOOL(pat):
               return findOBJECT(pat, myDATAWARE['ATKTOOL'])
           

        def initRelationships(self, data):
            if self.trace:
                print ('Loading ATT&CK Relationships data..')
            if self.loadFromSpreadsheet:
                return self.loadRelationshipsFromSheet (self.fname, 'ATKRELS', data)
            else:
                return self.loadRelationshipsFromJSON ( data )

        def loadRelationshipsFromJSON(self, data ):
            
            ret = []
            countr=0
            self.relationships = self.all_attack['relationships']
            for x in self.relationships:               
                for rel in x['objects']:
                    ret.append(ATKRELATION (rel['created'] if 'created' in rel else None,
                                            rel['created_by_ref'] if 'created_by_ref' in rel else None,
                                            rel['id'] if 'id' in rel else None,
                                            rel['modified'] if 'modified' in rel else None,
                                            rel['relationship_type'] if 'relationship_type' in rel else None,
                                            rel['description'] if 'description' in rel else None, 
                                            rel['source_ref'] if 'source_ref' in rel else None, 
                                            rel['target_ref'] if 'target_ref' in rel else None ))               
                                        
                    countr = countr+1
                    src = self.find (rel['source_ref'], data)
                    if not(src):
                        if self.trace:
                            print ('loadRelationships(): Relation', countr, ': SRC', rel['source_ref'], 'not found.')
            
                    tgt = self.find (rel['target_ref'], data)
                    if not(tgt):
                        if self.trace:
                            print ('loadRelationships(): Relation', countr, ': TGT', rel['target_ref'], 'not found.')
                
                    if (src) and (tgt):
                        desc = rel['description'] if 'description' in rel else None
                        if (rel['relationship_type']=='uses'):
                            src.addUses(tgt, desc)
                        elif (rel['relationship_type']=='mitigates'):
                            src.addMitigates(tgt, desc)
                            tgt.addCOA(src, desc)
                        elif (rel['relationship_type']=='relates-to'):
                            src.addRelates (tgt, desc)
                        elif (rel['relationship_type']=='revoked-by'):
                            src.addRevokes(tgt, desc)
            return ret
            
        def loadRelationshipsFromSheet(self, fname, sname, data ):
          book = openpyxl.load_workbook(fname, data_only=True) 
          sheet = book[sname]

          ret = []         
          countr=0
          for rel in sheet.rows:
              
               ret.append(ATKRELATION (rel[0].value,
                                       rel[1].value,
                                       rel[2].value,
                                       rel[3].value,
                                       rel[4].value,
                                       rel[5].value, 
                                       rel[6].value, 
                                       rel[7].value ))       
              
               countr = countr+1
               src = self.find (rel[6].value, data)
       
               tgt = self.find (rel[7].value, data)
        
               if (src) and (tgt):
                  desc = rel[5].value
                  if (rel[4].value =='uses'):
                    src.addUses(tgt, desc)
                  elif (rel[4].value =='mitigates'):
                    src.addMitigates(tgt, desc)
                    tgt.addCOA(src, desc)
                  elif (rel[4].value =='relates-to'):
                    src.addRelates (tgt, desc)
                  elif (rel[4].value =='revoked-by'):
                    src.addRevokes(tgt, desc)

          del (ret[0])
          return ret
     

        def exportRelationships(self):
            ret = []
            if self.relationships:
              for rel in self.relationships:
                ret.append(ATKRELATION (rel['created'],
                                        rel['created_by_ref'],
                                        rel['id'],
                                        rel['modified'],
                                        rel['relationship'],
                                        rel['relationship_description'], 
                                        rel['source_object'], 
                                        rel['target_object'] ))
            
            return ret
   
